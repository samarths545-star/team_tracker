from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
import pandas as pd
import sqlite3
import os
import io
from datetime import datetime
from werkzeug.utils import secure_filename
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import traceback

from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = "supersecretkey"

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "/"

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
DB_PATH       = os.path.join(BASE_DIR, "database.db")
STATIC_PATH   = os.path.join(BASE_DIR, "static")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(STATIC_PATH, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# ───────────────────────────────────────────────────────
# DATABASE INIT
# ───────────────────────────────────────────────────────

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    c.execute("""
    CREATE TABLE IF NOT EXISTS analytics (
        id                      INTEGER PRIMARY KEY AUTOINCREMENT,
        employee                TEXT,
        date                    TEXT,
        upload_label            TEXT DEFAULT '',
        total_calls             INTEGER DEFAULT 0,
        call_minutes            REAL    DEFAULT 0,
        total_faxes             INTEGER DEFAULT 0,
        fax_minutes             REAL    DEFAULT 0,
        cases                   INTEGER DEFAULT 0,
        facilities_total        INTEGER DEFAULT 0,
        records_received        INTEGER DEFAULT 0,
        expected_records        INTEGER DEFAULT 0,
        shall_be_receiving      INTEGER DEFAULT 0,
        correspondence_received INTEGER DEFAULT 0,
        summons_efile           INTEGER DEFAULT 0,
        summons_served          INTEGER DEFAULT 0,
        denials_received        INTEGER DEFAULT 0,
        cases_for_records       INTEGER DEFAULT 0,
        cases_for_summons       INTEGER DEFAULT 0
    )""")

    for tbl in ("call_log_rows", "fax_log_rows", "consolidated_rows"):
        c.execute(f"""
        CREATE TABLE IF NOT EXISTS {tbl} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee TEXT,
            row_hash TEXT UNIQUE
        )""")

    existing = [r[1] for r in c.execute("PRAGMA table_info(analytics)").fetchall()]
    for col, defn in [("upload_label","TEXT DEFAULT ''"),
                      ("cases_for_records","INTEGER DEFAULT 0"),
                      ("cases_for_summons","INTEGER DEFAULT 0")]:
        if col not in existing:
            c.execute(f"ALTER TABLE analytics ADD COLUMN {col} {defn}")

    conn.commit()
    conn.close()

init_db()

# ───────────────────────────────────────────────────────
# USERS
# ───────────────────────────────────────────────────────

USERS = {
    "Samarth":      {"password": "samarth1511",    "display": "Samarth"},
    "BigBossSteve": {"password": "masterlogin3217", "display": "Steve"},
    "Pragati":      {"password": "pragati1711",     "display": "Pragati"},
}
EMPLOYEES = ["Kavish", "Chirag", "Sahil", "Tushar"]

class User(UserMixin):
    def __init__(self, id):
        self.id      = id
        self.display = USERS[id]["display"]

@login_manager.user_loader
def load_user(uid):
    return User(uid) if uid in USERS else None

# ───────────────────────────────────────────────────────
# HELPERS
# ───────────────────────────────────────────────────────

def safe_numeric(val):
    if pd.isna(val) or val == "": return 0
    try: return float(val)
    except: return 0

def find_column(df, keywords):
    cols_lower = df.columns.str.lower().str.strip()
    for kw in (keywords if isinstance(keywords, list) else [keywords]):
        for idx, col in enumerate(cols_lower):
            if kw.lower() in col:
                return df.columns[idx]
    return None

def hash_row(row):
    return str(hash(tuple(str(v) for v in row.values)))

def filter_new_rows(df, employee, table_name, conn):
    cursor = conn.cursor()
    new_rows, new_hashes = [], []
    for _, row in df.iterrows():
        h = hash_row(row)
        cursor.execute(f"SELECT 1 FROM {table_name} WHERE row_hash=? AND employee=?", (h, employee))
        if cursor.fetchone() is None:
            new_rows.append(row)
            new_hashes.append(h)
    for h in new_hashes:
        try:
            cursor.execute(f"INSERT INTO {table_name} (employee,row_hash) VALUES (?,?)", (employee, h))
        except: pass
    conn.commit()
    return pd.DataFrame(new_rows, columns=df.columns) if new_rows else pd.DataFrame(columns=df.columns)

def get_filtered_df(date_from=None, date_to=None):
    conn   = sqlite3.connect(DB_PATH)
    query  = "SELECT * FROM analytics WHERE 1=1"
    params = []
    if date_from: query += " AND date >= ?"; params.append(date_from)
    if date_to:   query += " AND date <= ?"; params.append(date_to)
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df

# ───────────────────────────────────────────────────────
# FILE PROCESSING
# ───────────────────────────────────────────────────────

def process_consolidated(filepath, employee, upload_label=""):
    conn   = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    try:
        df = pd.read_excel(filepath, sheet_name=0)
        df.columns = df.columns.str.lower().str.strip()
        df = df[~df.iloc[:,0].astype(str).str.lower().str.contains('total', na=False)]

        new_df = filter_new_rows(df, employee, "consolidated_rows", conn)
        if new_df.empty:
            conn.close(); return 0

        cases_col          = find_column(new_df, ['cases','case'])
        facilities_col     = find_column(new_df, ['facilities','total'])
        records_col        = find_column(new_df, ['records received','records'])
        expected_col       = find_column(new_df, ['expected'])
        shall_col          = find_column(new_df, ['shall','receiving'])
        correspondence_col = find_column(new_df, ['correspondence','death'])
        efile_col          = find_column(new_df, ['efile','e-file','e_file'])
        served_col         = find_column(new_df, ['served','summons'])
        denials_col        = find_column(new_df, ['denial','denials'])
        cases_records_col  = find_column(new_df, ['cases for records','cases_records','record cases'])
        cases_summons_col  = find_column(new_df, ['cases for summons','cases_summons','summons cases'])

        s = {k:0 for k in ['cases','fac','rec','exp','shall','corr','efile','served','den','crec','csum']}
        for _, row in new_df.iterrows():
            if cases_col:          s['cases']  += int(safe_numeric(row[cases_col]))
            if facilities_col:     s['fac']    += int(safe_numeric(row[facilities_col]))
            if records_col:        s['rec']    += int(safe_numeric(row[records_col]))
            if expected_col:       s['exp']    += int(safe_numeric(row[expected_col]))
            if shall_col:          s['shall']  += int(safe_numeric(row[shall_col]))
            if correspondence_col: s['corr']   += int(safe_numeric(row[correspondence_col]))
            if efile_col:          s['efile']  += int(safe_numeric(row[efile_col]))
            if served_col:         s['served'] += int(safe_numeric(row[served_col]))
            if denials_col:        s['den']    += int(safe_numeric(row[denials_col]))
            if cases_records_col:  s['crec']   += int(safe_numeric(row[cases_records_col]))
            if cases_summons_col:  s['csum']   += int(safe_numeric(row[cases_summons_col]))

        if not cases_records_col: s['crec'] = s['cases']
        if not cases_summons_col: s['csum'] = s['efile'] + s['served']

        date_val = datetime.now().strftime("%Y-%m-%d")
        cursor.execute("""
            INSERT INTO analytics
            (employee,date,upload_label,cases,facilities_total,records_received,
             expected_records,shall_be_receiving,correspondence_received,
             summons_efile,summons_served,denials_received,
             cases_for_records,cases_for_summons)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (employee,date_val,upload_label,
             s['cases'],s['fac'],s['rec'],s['exp'],s['shall'],s['corr'],
             s['efile'],s['served'],s['den'],s['crec'],s['csum']))
        conn.commit()
        return len(new_df)
    except Exception as e:
        print(f"Consolidated error: {e}"); traceback.print_exc(); return 0
    finally:
        conn.close()


def process_calls(filepath, employee, upload_label=""):
    conn   = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    try:
        df = pd.read_csv(filepath) if filepath.endswith('.csv') else pd.read_excel(filepath)
        df.columns = df.columns.str.lower().str.strip()
        dur_col = find_column(df, ['duration'])
        if dur_col is None:
            conn.close(); return 0

        new_df = filter_new_rows(df, employee, "call_log_rows", conn)
        if new_df.empty:
            conn.close(); return 0

        total_mins = 0
        for _, row in new_df.iterrows():
            try:
                parts = str(row[dur_col]).strip().split(':')
                if len(parts)==3: total_mins += int(parts[0])*60+int(parts[1])+int(parts[2])/60
                elif len(parts)==2: total_mins += int(parts[0])+int(parts[1])/60
            except: pass

        date_val = datetime.now().strftime("%Y-%m-%d")
        cursor.execute(
            "SELECT id FROM analytics WHERE employee=? AND date=? AND upload_label=? AND total_calls=0 AND records_received=0",
            (employee,date_val,upload_label))
        res = cursor.fetchone()
        if res:
            cursor.execute("UPDATE analytics SET total_calls=total_calls+?,call_minutes=call_minutes+? WHERE id=?",
                           (len(new_df),total_mins,res[0]))
        else:
            cursor.execute("INSERT INTO analytics(employee,date,upload_label,total_calls,call_minutes) VALUES(?,?,?,?,?)",
                           (employee,date_val,upload_label,len(new_df),total_mins))
        conn.commit(); return len(new_df)
    except Exception as e:
        print(f"Calls error: {e}"); traceback.print_exc(); return 0
    finally:
        conn.close()


def process_faxes(filepath, employee, upload_label=""):
    conn   = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    try:
        df = pd.read_csv(filepath) if filepath.endswith('.csv') else pd.read_excel(filepath)
        df.columns = df.columns.str.lower().str.strip()
        # Drop exact duplicate rows within the sheet (same fax number/extension etc.)
        df = df.drop_duplicates()

        new_df = filter_new_rows(df, employee, "fax_log_rows", conn)
        if new_df.empty:
            conn.close(); return 0

        total_faxes = len(new_df)
        fax_minutes = total_faxes * 20
        date_val = datetime.now().strftime("%Y-%m-%d")

        cursor.execute(
            "SELECT id FROM analytics WHERE employee=? AND date=? AND upload_label=? AND total_faxes=0 AND records_received=0",
            (employee,date_val,upload_label))
        res = cursor.fetchone()
        if res:
            cursor.execute("UPDATE analytics SET total_faxes=total_faxes+?,fax_minutes=fax_minutes+? WHERE id=?",
                           (total_faxes,fax_minutes,res[0]))
        else:
            cursor.execute("INSERT INTO analytics(employee,date,upload_label,total_faxes,fax_minutes) VALUES(?,?,?,?,?)",
                           (employee,date_val,upload_label,total_faxes,fax_minutes))
        conn.commit(); return total_faxes
    except Exception as e:
        print(f"Faxes error: {e}"); traceback.print_exc(); return 0
    finally:
        conn.close()

# ───────────────────────────────────────────────────────
# SCORING  (all out of 100)
# ───────────────────────────────────────────────────────
#
#  Communication (/35) = (call_min + fax_min) / team_max  × 35
#  Records       (/40) = min(received / expected, 1.0)    × 40
#  Legal         (/25) = (e-files + served)  / team_max   × 25
#  Final                = Communication + Records + Legal   [max 100]
#  A+ ≥90  A ≥80  B ≥70  C ≥60  D <60

def calculate_scores(df):
    df = df.copy()
    df["raw_comm"]  = df["call_minutes"] + df["fax_minutes"]
    max_comm        = df["raw_comm"].max() or 1
    df["communication_score"] = (df["raw_comm"] / max_comm * 35).round(2)

    max_rec = df["records_received"].max() or 1
    def rec_score(row):
        if row["expected_records"] > 0:
            return min(row["records_received"]/row["expected_records"], 1.0) * 40
        return (row["records_received"]/max_rec) * 40
    df["records_score"] = df.apply(rec_score, axis=1).round(2)

    df["raw_legal"] = df["summons_efile"] + df["summons_served"]
    max_legal       = df["raw_legal"].max() or 1
    df["legal_score"] = (df["raw_legal"] / max_legal * 25).round(2)

    df["final_score"] = (df["communication_score"]+df["records_score"]+df["legal_score"]).round(2)

    def grade(s):
        if s>=90: return "A+"
        if s>=80: return "A"
        if s>=70: return "B"
        if s>=60: return "C"
        return "D"
    df["Grade"] = df["final_score"].apply(grade)
    return df

def generate_insights(df):
    insights = {}
    avg_comm  = df["communication_score"].mean() or 0
    avg_rec   = df["records_score"].mean() or 0
    avg_legal = df["legal_score"].mean() or 0
    avg_final = df["final_score"].mean() or 0

    for _, row in df.iterrows():
        emp = row["employee"]
        S, I, G = [], [], []   # strengths, improvements, suggestions

        # Communication
        if row["communication_score"] >= avg_comm*1.1:
            S.append("Strong outreach — calls & faxes above team average")
        elif row["communication_score"] < avg_comm*0.8 and avg_comm>0:
            I.append("Communication time below team average")
            G.append("Increase daily call volume and fax outreach")
        if row["total_calls"]>0 and (row["call_minutes"]/row["total_calls"])<2:
            G.append("Calls average under 2 min — aim for more substantive conversations")

        # Records
        if row["records_score"] >= avg_rec*1.1:
            S.append("Excellent medical records retrieval")
        elif row["records_score"] < avg_rec*0.8 and avg_rec>0:
            I.append("Records retrieval below team average")
        if row["expected_records"]>0:
            pct = (row["records_received"]/row["expected_records"])*100
            if pct>=90:
                S.append(f"{pct:.0f}% of expected records received — near perfect")
            elif pct<60:
                I.append(f"Only {pct:.0f}% of expected records received — {int(row['expected_records']-row['records_received'])} still pending")
                G.append("Follow up aggressively on outstanding record requests")

        # Legal
        if row["legal_score"] >= avg_legal*1.1:
            S.append("High legal activity — strong e-filing & serving")
        elif row["legal_score"] < avg_legal*0.8 and avg_legal>0:
            I.append("Legal activity (summons) below team average")
            G.append("Prioritise e-filing and serving pending summons")
        if row["summons_efile"]>0 and row["summons_served"]==0:
            G.append("E-files submitted but none served yet — follow up on serving")

        # Denials
        if row.get("denials_received",0)>0:
            I.append(f"{int(row['denials_received'])} denial(s) received — review and appeal where applicable")

        if row["final_score"]   >= avg_final*1.15:
            overall = "🌟 <strong>Top performer</strong> — excellent work across all categories."
        elif row["final_score"] >= avg_final*0.9:
            overall = "✅ <strong>On track</strong> — performing at team average, keep it consistent."
        else:
            overall = "⚠️ <strong>Needs attention</strong> — key metrics below team average."

        parts = [overall]
        if S: parts.append("💪 <em>Strengths:</em> " + "; ".join(S) + ".")
        if I: parts.append("📉 <em>Improve:</em> "   + "; ".join(I) + ".")
        if G: parts.append("💡 <em>Suggestions:</em> " + "; ".join(G) + ".")
        insights[emp] = " ".join(parts)
    return insights

# ───────────────────────────────────────────────────────
# ROUTES
# ───────────────────────────────────────────────────────

@app.route("/", methods=["GET","POST"])
def login():
    if request.method == "POST":
        u, p = request.form.get("username",""), request.form.get("password","")
        if u in USERS and USERS[u]["password"]==p:
            login_user(User(u)); return redirect("/dashboard")
        flash("Invalid username or password")
    return render_template("login.html")


@app.route("/dashboard")
@login_required
def dashboard():
    date_from = request.args.get("date_from","")
    date_to   = request.args.get("date_to","")

    conn = sqlite3.connect(DB_PATH)
    history_df = pd.read_sql_query(
        "SELECT id,employee,date,upload_label,total_calls,total_faxes,records_received FROM analytics ORDER BY date DESC,id DESC", conn)
    conn.close()

    raw = get_filtered_df(date_from or None, date_to or None)

    formula = (
        "<b>Scoring Formula (max 100):</b><br>"
        "• <b>Communication /35</b> = (call_minutes + fax_minutes) ÷ team_max × 35<br>"
        "• <b>Records /40</b> = min(records_received ÷ expected_records, 1.0) × 40<br>"
        "• <b>Legal /25</b> = (summons_efile + summons_served) ÷ team_max × 25<br>"
        "• Grades: A+ ≥90 | A ≥80 | B ≥70 | C ≥60 | D &lt;60"
    )

    if raw.empty:
        return render_template("master_dashboard.html",
            employees=EMPLOYEES, data=[], user=current_user.display,
            insights={}, date_from=date_from, date_to=date_to,
            history=history_df.to_dict(orient="records"), formula=formula)

    for col in ["cases_for_records","cases_for_summons"]:
        if col not in raw.columns: raw[col]=0

    df = raw.groupby("employee").sum(numeric_only=True).reset_index()
    df = calculate_scores(df)
    df = df.sort_values("final_score", ascending=False)
    df["Rank"] = range(1, len(df)+1)
    insights = generate_insights(df)

    return render_template("master_dashboard.html",
        employees=EMPLOYEES, data=df.to_dict(orient="records"),
        user=current_user.display, insights=insights,
        date_from=date_from, date_to=date_to,
        history=history_df.to_dict(orient="records"), formula=formula)


@app.route("/upload", methods=["POST"])
@login_required
def upload_file():
    employee     = request.form.get("employee","")
    upload_label = request.form.get("upload_label","") or datetime.now().strftime("Week of %d %b %Y")

    if not employee or employee not in EMPLOYEES:
        flash("Please select a valid employee"); return redirect(url_for("dashboard"))

    try:
        for key, processor in [("consolidated", process_consolidated),
                                ("calls",        process_calls),
                                ("faxes",        process_faxes)]:
            if key in request.files and request.files[key].filename:
                f  = request.files[key]
                fp = os.path.join(app.config["UPLOAD_FOLDER"], secure_filename(f.filename))
                f.save(fp)
                n  = processor(fp, employee, upload_label)
                os.remove(fp)
                label = key.capitalize()
                flash(f"✓ {label}: {n} new rows added for {employee}" if n else
                      f"ℹ️ {label} for {employee}: no new data (all duplicates)")
    except Exception as e:
        flash(f"Error: {str(e)}"); print(traceback.format_exc())

    return redirect(url_for("dashboard"))


@app.route("/delete_upload/<int:row_id>", methods=["POST"])
@login_required
def delete_upload(row_id):
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DELETE FROM analytics WHERE id=?", (row_id,))
    conn.commit(); conn.close()
    flash("✓ Upload entry deleted. Dashboard reflects the remaining data.")
    return redirect(url_for("dashboard"))


# ── PDF Export ─────────────────────────────

@app.route("/export/pdf")
@login_required
def export_pdf():
    date_from = request.args.get("date_from","")
    date_to   = request.args.get("date_to","")
    raw = get_filtered_df(date_from or None, date_to or None)
    if raw.empty:
        flash("No data to export."); return redirect(url_for("dashboard"))

    for col in ["cases_for_records","cases_for_summons"]:
        if col not in raw.columns: raw[col]=0

    df = raw.groupby("employee").sum(numeric_only=True).reset_index()
    df = calculate_scores(df)
    df = df.sort_values("final_score", ascending=False)
    df["Rank"] = range(1, len(df)+1)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=0.4*inch, rightMargin=0.4*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story  = []
    story.append(Paragraph("Performance Analytics Report", styles["Title"]))
    story.append(Paragraph(
        f"Period: {date_from or 'All time'} → {date_to or 'Today'}   |   "
        f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", styles["Normal"]))
    story.append(Spacer(1,12))

    hdrs = ["Rank","Employee","Calls","Call Min","Faxes","Fax Min",
            "Records","Expected","E-Files","Served",
            "Comm/35","Rec/40","Legal/25","Score/100","Grade"]
    rows = [hdrs]
    for _, row in df.iterrows():
        rows.append([int(row["Rank"]), row["employee"],
            int(row["total_calls"]), round(row["call_minutes"],1),
            int(row["total_faxes"]), round(row["fax_minutes"],1),
            int(row["records_received"]), int(row["expected_records"]),
            int(row["summons_efile"]), int(row["summons_served"]),
            round(row["communication_score"],1), round(row["records_score"],1),
            round(row["legal_score"],1), round(row["final_score"],2), row["Grade"]])

    cw = [0.38*inch,1.0*inch,0.48*inch,0.58*inch,0.48*inch,0.58*inch,
          0.6*inch,0.68*inch,0.52*inch,0.52*inch,
          0.65*inch,0.65*inch,0.65*inch,0.72*inch,0.45*inch]
    t = Table(rows, colWidths=cw, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F3C88")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#EEF2FF")]),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#CCCCCC")),
        ("ROWHEIGHT",(0,0),(-1,-1),16),
    ]))
    story.append(t)
    story.append(Spacer(1,14))
    story.append(Paragraph("<b>Formula:</b> Comm(/35)=(call_min+fax_min)÷max×35 | "
                           "Records(/40)=min(received÷expected,1)×40 | "
                           "Legal(/25)=(efile+served)÷max×25 | Total=max 100",
                           styles["Normal"]))
    doc.build(story)
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf", as_attachment=True,
                     download_name=f"performance_{datetime.now().strftime('%Y%m%d')}.pdf")


# ── Excel Export ───────────────────────────

@app.route("/export/excel")
@login_required
def export_excel():
    date_from = request.args.get("date_from","")
    date_to   = request.args.get("date_to","")
    raw = get_filtered_df(date_from or None, date_to or None)
    if raw.empty:
        flash("No data to export."); return redirect(url_for("dashboard"))

    for col in ["cases_for_records","cases_for_summons"]:
        if col not in raw.columns: raw[col]=0

    df = raw.groupby("employee").sum(numeric_only=True).reset_index()
    df = calculate_scores(df)
    df = df.sort_values("final_score", ascending=False)
    df["Rank"] = range(1, len(df)+1)

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Performance Report"

    hdr_fill  = PatternFill("solid", fgColor="1F3C88")
    hdr_font  = Font(color="FFFFFF", bold=True, size=11)
    alt_fill  = PatternFill("solid", fgColor="EEF2FF")
    center    = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="CCCCCC")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    grade_clr = {"A+":"92D050","A":"00B0F0","B":"00B0A0","C":"FFBF00","D":"FF5050"}

    headers = ["Rank","Employee","Total Calls","Call Minutes","Total Faxes","Fax Minutes",
               "Records Received","Expected Records","Cases (Records)","Facilities Total",
               "Cases (Summons)","E-Files","Served",
               "Communication /35","Records /40","Legal /25","Final Score /100","Grade"]
    cols_map = ["Rank","employee","total_calls","call_minutes","total_faxes","fax_minutes",
                "records_received","expected_records","cases_for_records","facilities_total",
                "cases_for_summons","summons_efile","summons_served",
                "communication_score","records_score","legal_score","final_score","Grade"]

    ws.append(headers)
    for cell in ws[1]:
        cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=bdr

    for i,(_, row) in enumerate(df.iterrows(), start=2):
        data = []
        for col in cols_map:
            v = row.get(col,0)
            if isinstance(v, float): v = round(v,2)
            data.append(v)
        ws.append(data)
        fill = PatternFill("solid",fgColor="FFFFFF") if i%2==0 else alt_fill
        for j, cell in enumerate(ws[i], start=1):
            cell.alignment=center; cell.border=bdr; cell.fill=fill
            if headers[j-1]=="Grade":
                g = row.get("Grade","D")
                cell.fill=PatternFill("solid",fgColor=grade_clr.get(g,"FF5050"))
                cell.font=Font(bold=True)

    for ci, col_cells in enumerate(ws.columns, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = min(
            max(len(str(c.value or "")) for c in col_cells)+3, 22)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Scoring Formula")
    for r in [
        ["Performance Scoring Formula"],[""],
        ["Component","Max Points","Formula"],
        ["Communication",35,"(call_minutes + fax_minutes) ÷ team_max × 35"],
        ["Records",40,"min(records_received ÷ expected_records, 1.0) × 40"],
        ["Legal",25,"(summons_efile + summons_served) ÷ team_max × 25"],
        ["TOTAL",100,"Communication + Records + Legal"],[""],
        ["Grade","Min Score",""],["A+",90,""],["A",80,""],["B",70,""],["C",60,""],["D","<60",""]
    ]: ws2.append(r)
    ws2["A1"].font = Font(bold=True, size=13)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"performance_{datetime.now().strftime('%Y%m%d')}.xlsx")


@app.route("/forgot_password", methods=["GET","POST"])
def forgot_password():
    return render_template("forgot_password.html")

@app.route("/reset_password", methods=["GET","POST"])
def reset_password():
    return render_template("reset_password.html")

@app.route("/logout")
@login_required
def logout():
    logout_user(); return redirect("/")

if __name__ == "__main__":
    app.run(debug=True)
